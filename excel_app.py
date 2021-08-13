
from flask import Flask, render_template, url_for, request, jsonify, current_app, make_response, send_from_directory, send_file
from openpyxl import load_workbook
import os

app = Flask(__name__)




@app.route('/')
def index():
    return render_template('index.html')


@app.route('/process_qtc', methods=['POST', 'GET'])
def process_qt_calculation():
    if request.method == "POST":
        data = request.get_json()

        SITE_ROOT = os.path.realpath(os.path.dirname(__file__))

        wb = load_workbook(os.path.join(SITE_ROOT, "static", "maurizio_montanari_excel.xlsx"))
        ws = wb.worksheets[0]
        mx = ws.max_row + 1
        ws['A' + str(mx)] = data["Cliente"]
        ws['B' + str(mx)] = data["Nome_compilatore"]
        ws['C' + str(mx)] = data["Cognome_compilatore"]
        ws['D' + str(mx)] = data["Orario"]
        ws['E' + str(mx)] = data["Data"]
        ws['F' + str(mx)] = data["Tipo_operazione"]
        ws['G' + str(mx)] = data["Descrizione_post"]
        ws['H' + str(mx)] = data["Testo_post"]
        ws['I' + str(mx)] = data["Hashtag"]

        wb.save(os.path.join(SITE_ROOT, "static", "maurizio_montanari_excel.xlsx"))


    results = {'processed': 'true'}
    return jsonify(results)


@app.route('/download',  methods=['GET', 'POST'])
def download():
    SITE_ROOT = os.path.realpath(os.path.dirname(__file__))
    return send_file(os.path.join(SITE_ROOT, "static", "maurizio_montanari_excel.xlsx"), as_attachment=True)



if __name__ == "__main__":
    app.run()
